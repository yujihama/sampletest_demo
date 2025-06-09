"""
Excel入力欄特定とJSON化
"""

import os
import json
import base64
import logging
import tempfile
from pathlib import Path
from typing import Dict, List, Literal, Optional, TypedDict

# LangChain関連のインポート
from langchain_core.messages import HumanMessage
from langchain_openai import ChatOpenAI
from langgraph.graph import StateGraph, END
from pydantic import BaseModel, Field

# Excel操作関連のインポート
import openpyxl
from openpyxl.styles import PatternFill
import subprocess

# 環境変数の読み込み
from dotenv import load_dotenv
load_dotenv()

logger = logging.getLogger(__name__)

# Pydanticモデル: 入力欄情報
class ExcelField(BaseModel):
    """Excelの入力欄情報を表すモデル"""
    cell_id: str = Field(..., description="セル番号（例: A1, B2）")
    description: str = Field(..., description="そのセルに記入すべき内容の説明")

class ExcelFormFields(BaseModel):
    """Excelフォームの入力欄情報のコレクション"""
    fields: List[ExcelField] = Field(..., description="検出された入力欄のリスト")
    reason: str = Field(..., description="判断根拠")
    
class CollectExcelFormFields(BaseModel):
    """Excelフォームの入力欄情報の修正箇所"""
    add_fields: List[ExcelField] = Field(..., description="追加する入力欄のリスト")
    delete_fields: List[ExcelField] = Field(..., description="削除する入力欄のリスト")
    reason: str = Field(..., description="判断根拠")

# Pydanticモデル: 検証結果
class ValidationResult(BaseModel):
    """入力欄の検証結果を表すモデル"""
    status: Literal["OK", "修正が必要"] = Field(..., description="検証結果のステータス")
    issues: Optional[List[str]] = Field(None, description="問題点のリスト（ステータスが「修正が必要」の場合）")
    suggestions: Optional[List[str]] = Field(None, description="修正提案のリスト（ステータスが「修正が必要」の場合）")

# 状態の型定義
class ExcelFormState(TypedDict):
    excel_file: str
    output_dir: Optional[Path]
    max_iterations: int
    current_iteration: int
    extracted_text_file: str
    original_excel_capture: str
    estimated_fields: Dict[str, str] 
    structured_fields: ExcelFormFields
    highlighted_excel: str
    highlighted_captures: List[str]
    validation_result: str 
    structured_validation: ValidationResult
    validation_status: Literal["OK", "修正が必要", "エラー"]
    final_json: str
    status: Literal["進行中", "完了", "エラー"]
    error_message: str
    temp_excel_for_capture: str

# 1. Excelデータのテキスト化と画像キャプチャ
def extract_excel_data_and_capture(state: ExcelFormState) -> ExcelFormState:
    """
    Excelファイルからテキストデータを抽出し、画像キャプチャを取得する
    """
    logger.info(f"Excelテキスト抽出と画像キャプチャ開始: {state['excel_file']}")
    temp_excel_file_for_capture_path = None # finallyで使うため、ここで定義

    try:
        # 実際の保存先ベースディレクトリを決定
        user_defined_output_dir = state.get("output_dir")
        if user_defined_output_dir and str(user_defined_output_dir).strip():
            base_save_path = Path(user_defined_output_dir)
        else:
            base_save_path = Path(state["excel_file"]).parent
        
        final_output_dir = base_save_path / "format_data"

        # format_data ディレクトリが存在すれば削除して再作成
        if final_output_dir.exists():
            import shutil # shutilをインポート
            shutil.rmtree(final_output_dir)
            logger.info(f"既存の出力ディレクトリ {final_output_dir} を削除しました。")
        
        final_output_dir.mkdir(exist_ok=True, parents=True)
        
        # キャプチャ用ディレクトリ作成
        captures_dir = final_output_dir / "captures"
        captures_dir.mkdir(exist_ok=True, parents=True)

        # キャプチャ前に既存のPNGファイルを削除 (format_dataを削除するため、この処理は実質不要になるが、残しても問題はない)
        # for png_file in captures_dir.glob("*.png"):
        #     try:
        #         png_file.unlink()
        #         logger.info(f"既存のキャプチャファイルを削除: {png_file}")
        #     except Exception as e:
        #         logger.warning(f"キャプチャファイルの削除に失敗: {png_file} ({e})")
        
        # Excelファイルを開く (テキスト抽出用)
        workbook_orig = openpyxl.load_workbook(state["excel_file"])
        
        # キャプチャ用にExcelを準備 (印刷範囲設定)
        workbook_for_capture = openpyxl.load_workbook(state["excel_file"])
        
        try:
            for sheet_capture in workbook_for_capture:
                try:
                    dimension = sheet_capture.calculate_dimension()
                    if dimension:
                        sheet_capture.print_area = dimension
                        # LibreOffice が印刷範囲を確実に認識するよう fitToPage を有効化
                        sheet_capture.page_setup.fitToPage = True
                        logger.info(f"シート '{sheet_capture.title}' の印刷範囲を {dimension} に設定しました")
                except Exception as e_dim:
                    logger.warning(f"シート '{sheet_capture.title}' の印刷範囲設定エラー: {e_dim}")

            # 印刷範囲設定済みのExcelを一時ファイルに保存
            # delete=False にして、sofficeがファイルを使用後に手動で削除
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", prefix="capture_") as tmp_excel_file:
                workbook_for_capture.save(tmp_excel_file.name)
                temp_excel_file_for_capture_path = tmp_excel_file.name # finally節で使うためにパスを保存
            
            logger.info(f"印刷範囲設定済みのExcelを一時ファイル '{temp_excel_file_for_capture_path}' に保存しました。")

        except Exception as e_save:
            logger.error(f"一時Excelファイルの保存中にエラーが発生しました: {e_save}")
            raise 

        # 抽出結果を格納するテキスト
        extracted_text = ""
        
        # 各シートの処理 (workbook_orig を使用)
        for sheet_name in workbook_orig.sheetnames:
            sheet = workbook_orig[sheet_name]
            
            # シート名の追加
            extracted_text += f"## シート名: {sheet_name}\n"
            
            # 結合セル情報の抽出
            merged_cells = []
            for merged_cell_range in sheet.merged_cells.ranges:
                merged_cells.append(str(merged_cell_range))
            
            if merged_cells:
                extracted_text += "### 結合セル情報:\n"
                for cell_range in merged_cells:
                    extracted_text += f"- {cell_range}\n"
            
            # セルデータの抽出
            extracted_text += "### セルデータ:\n"
            extracted_text += "| セル | 値 | 書式 |\n"
            extracted_text += "|-----|----|--------|\n"
            
            for row in sheet.iter_rows():
                for cell in row:
                    # セルが空でない場合のみ処理
                    if cell.value is not None:
                        cell_addr = f"{cell.column_letter}{cell.row}"
                        cell_value = str(cell.value)
                        
                        # 書式情報の取得
                        format_info = []
                        if cell.font.bold:
                            format_info.append("太字")
                        if cell.fill.fill_type == "solid":
                            fill_color = cell.fill.start_color.index
                            if fill_color != "00000000":  # デフォルト色でない場合
                                format_info.append(f"背景色:{fill_color}")
                        
                        format_str = ", ".join(format_info) if format_info else "-"
                        
                        # テーブルに行を追加
                        extracted_text += f"| {cell_addr} | {cell_value} | {format_str} |\n"
        
        # 抽出結果をファイルに保存
        extracted_text_file = final_output_dir / "extracted_excel_text.md"
        with open(extracted_text_file, "w", encoding="utf-8") as f:
            f.write(extracted_text)
        
        logger.info(f"Excelテキスト抽出完了: {extracted_text_file}")
        
        original_capture_path = None
        if temp_excel_file_for_capture_path and os.path.exists(temp_excel_file_for_capture_path):
            command = f"soffice --headless --convert-to png \"{str(temp_excel_file_for_capture_path)}\" --outdir \"{str(captures_dir)}\""
            logger.info(f"実行コマンド: {command}")
            subprocess.run(command, shell=True, check=True)
            
            temp_excel_basename = os.path.splitext(os.path.basename(temp_excel_file_for_capture_path))[0]
            expected_capture_name = f"{temp_excel_basename}.png"
            generated_capture_path = captures_dir / expected_capture_name

            if not generated_capture_path.exists():
                png_files_in_captures_dir = list(captures_dir.glob(f"{temp_excel_basename}*.png")) 
                if png_files_in_captures_dir:
                    generated_capture_path = png_files_in_captures_dir[0]
                    logger.info(f"期待された名前 {expected_capture_name} が見つからず、代わりに {generated_capture_path.name} を使用します。")
                else: 
                    logger.warning(f"キャプチャファイル {expected_capture_name} または {temp_excel_basename}*.png が見つかりません。captures_dir: {captures_dir}")

            if generated_capture_path and generated_capture_path.exists():
                final_capture_name = "original_excel.png"
                original_capture_path = captures_dir / final_capture_name
                if generated_capture_path != original_capture_path: 
                     if original_capture_path.exists(): 
                        original_capture_path.unlink()
                     os.rename(generated_capture_path, original_capture_path)
                logger.info(f"元Excelのキャプチャ完了: {original_capture_path}")
            else:
                logger.error(f"SofficeによるPNG変換後、キャプチャファイルが見つかりませんでした。一時ファイル: {temp_excel_file_for_capture_path}")
        else:
            logger.warning("印刷範囲設定済みの一時Excelファイルが見つからないため、キャプチャをスキップします。")
        
        return {
            **state,
            "extracted_text_file": str(extracted_text_file),
            "original_excel_capture": str(original_capture_path) if original_capture_path else "", 
            "status": "進行中"
        }
        
    except Exception as e:
        logger.error(f"Excelテキスト抽出と画像キャプチャエラー: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return {
            **state,
            "status": "エラー",
            "error_message": f"Excelテキスト抽出と画像キャプチャエラー: {str(e)}\n{traceback.format_exc()}"
        }
    finally:
        if temp_excel_file_for_capture_path and os.path.exists(temp_excel_file_for_capture_path):
            try:
                os.remove(temp_excel_file_for_capture_path)
                logger.info(f"一時ファイル '{temp_excel_file_for_capture_path}' を削除しました。")
            except Exception as e_remove:
                logger.warning(f"一時ファイル '{temp_excel_file_for_capture_path}' の削除に失敗しました: {e_remove}")

# 2. マルチモーダルLLMによる入力欄の推定（structured_output使用）
def estimate_fields_with_multimodal_llm(state: ExcelFormState) -> ExcelFormState:
    """
    テキストデータと画像キャプチャを使用してマルチモーダルLLMで入力欄を推定する
    structured_outputを使用して確実に指定の形式で結果を受け取る
    """
    logger.info(f"マルチモーダルLLMによる入力欄推定開始 (v{state['current_iteration']})")
    
    try:
        # 抽出されたテキストを読み込む
        with open(state["extracted_text_file"], "r", encoding="utf-8") as f:
            extracted_text = f.read()
        
        # 画像をbase64エンコード
        with open(state["original_excel_capture"], "rb") as img_file:
            base64_image = base64.b64encode(img_file.read()).decode("utf-8")
        
        # マルチモーダルLLMクライアントの初期化（structured_output使用）
        llm = ChatOpenAI(
            model="gpt-4.1-mini",
            temperature=0
        ).with_structured_output(ExcelFormFields)
        
        # プロンプトの作成
        prompt = f"""
あなたはExcelフォームの入力欄を特定する専門家です。

以下はExcelファイルから抽出したテキスト情報と、そのExcelシートの画像です。
このExcelファイルは入力フォームであり、ユーザーが情報を入力するセルを特定してください。

テキスト情報:
{extracted_text}

入力欄の特徴：
- 空白セル
- ラベル（太字や背景色付きのセル）の隣や下にある空白セル
- 表形式の場合、ヘッダー行の下の空白セル
- 既に値が入力されているセルでも、それが例や初期値と思われる場合は入力欄として扱う

画像とテキスト情報の両方を参考にして、入力欄を特定してください。
"""
        
        # マルチモーダルLLMに問い合わせ
        response = llm.invoke([
            HumanMessage(content=[
                {"type": "text", "text": prompt},
                {
                    "type": "image_url",
                    "image_url": {
                        "url": f"data:image/png;base64,{base64_image}"
                    }
                }
            ]),
        ])
        
        # 構造化された応答を取得
        structured_fields = response
        
        # 従来の形式（Dict[str, str]）に変換（互換性のため）
        estimated_fields = {}
        for field in structured_fields.fields:
            estimated_fields[field.cell_id] = field.description
        
        # 実際の保存先ベースディレクトリを決定
        user_defined_output_dir = state.get("output_dir")
        if user_defined_output_dir and str(user_defined_output_dir).strip():
            base_save_path = Path(user_defined_output_dir)
        else:
            base_save_path = Path(state["excel_file"]).parent
        
        final_output_dir = base_save_path / "format_data"
        final_output_dir.mkdir(exist_ok=True, parents=True)
        
        # 構造化された形式を保存
        structured_fields_file = final_output_dir / f"structured_fields_v{state['current_iteration']}.json"
        with open(structured_fields_file, "w", encoding="utf-8") as f:
            f.write(structured_fields.model_dump_json(indent=2))
        
        # 従来の形式も保存（互換性のため）
        estimated_fields_file = final_output_dir / f"estimated_fields_v{state['current_iteration']}.json"
        with open(estimated_fields_file, "w", encoding="utf-8") as f:
            json.dump(estimated_fields, f, ensure_ascii=False, indent=2)
        
        logger.info(f"マルチモーダルLLMによる入力欄推定完了: {structured_fields_file}")
        
        # 状態の更新
        return {
            **state,
            "estimated_fields": estimated_fields,
            "structured_fields": structured_fields,
            "status": "進行中"
        }
        
    except Exception as e:
        logger.error(f"マルチモーダルLLMによる入力欄推定エラー: {str(e)}")
        return {
            **state,
            "status": "エラー",
            "error_message": f"マルチモーダルLLMによる入力欄推定エラー: {str(e)}"
        }

# 3. 入力欄のハイライト
def highlight_fields(state: ExcelFormState) -> ExcelFormState:
    """
    推定された入力欄をハイライトする
    """
    logger.info(f"入力欄のハイライト開始 (v{state['current_iteration']})")
    
    try:
        # 実際の保存先ベースディレクトリを決定
        user_defined_output_dir = state.get("output_dir")
        if user_defined_output_dir and str(user_defined_output_dir).strip():
            base_save_path = Path(user_defined_output_dir)
        else:
            base_save_path = Path(state["excel_file"]).parent
        
        final_output_dir = base_save_path / "format_data"
        final_output_dir.mkdir(exist_ok=True, parents=True)
        
        # 元のExcelファイルをコピー
        workbook = openpyxl.load_workbook(state["excel_file"])
        
        # 黄色のハイライト用フィル
        highlight_fill = PatternFill(
            start_color="FFFF00",
            end_color="FFFF00",
            fill_type="solid"
        )
        
        # 推定された入力欄をハイライト
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            for cell_addr in state["estimated_fields"].keys():
                try:
                    # セルアドレスが有効かチェック
                    if len(cell_addr) >= 2 and cell_addr[0].isalpha() and cell_addr[1:].isdigit():
                        cell = sheet[cell_addr]
                        original_value = cell.value # 元の値を取得
                        cell.fill = highlight_fill
                        if original_value is not None and str(original_value).strip() != "":
                            cell.value = f"{cell_addr}:{original_value}" # セルアドレスと元の値を連結
                        else:
                            cell.value = cell_addr # 元の値が空ならセルアドレスのみ設定
                except Exception as cell_error:
                    logger.warning(f"セル {cell_addr} のハイライトまたは値設定中にエラー: {str(cell_error)}")
        
        # ハイライト済みExcelを保存
        highlighted_excel = final_output_dir / f"highlighted_excel_v{state['current_iteration']}.xlsx"
        workbook.save(highlighted_excel)
        
        logger.info(f"入力欄のハイライト完了: {highlighted_excel}")
        
        # 状態の更新
        return {
            **state,
            "highlighted_excel": str(highlighted_excel),
            "status": "進行中"
        }
        
    except Exception as e:
        logger.error(f"入力欄のハイライトエラー: {str(e)}")
        return {
            **state,
            "status": "エラー",
            "error_message": f"入力欄のハイライトエラー: {str(e)}"
        }

# 4. ハイライト済みExcelのキャプチャ取得
def capture_highlighted_excel(state: ExcelFormState) -> ExcelFormState:
    """
    ハイライト済みExcelのキャプチャを取得する
    """
    logger.info(f"ハイライト済みExcelキャプチャ開始 (v{state['current_iteration']})")
    
    try:
        # 実際の保存先ベースディレクトリを決定
        user_defined_output_dir = state.get("output_dir")
        if user_defined_output_dir and str(user_defined_output_dir).strip():
            base_save_path = Path(user_defined_output_dir)
        else:
            base_save_path = Path(state["excel_file"]).parent
        
        final_output_dir = base_save_path / "format_data"
        final_output_dir.mkdir(exist_ok=True, parents=True)
        
        # キャプチャ用ディレクトリ作成
        captures_dir = final_output_dir / "captures"
        captures_dir.mkdir(exist_ok=True, parents=True)

        # キャプチャ前に既存のPNGファイルを削除
        # for png_file in captures_dir.glob("*.png"):
        #     try:
        #         png_file.unlink()
        #         logger.info(f"既存のキャプチャファイルを削除: {png_file}")
        #     except Exception as e:
        #         logger.warning(f"キャプチャファイルの削除に失敗: {png_file} ({e})")

        # ハイライト済みExcelファイルをロードし、印刷範囲を設定
        highlighted_excel_path_str = state["highlighted_excel"]
        workbook_hl = openpyxl.load_workbook(highlighted_excel_path_str)
        sheet_names_for_loop = list(workbook_hl.sheetnames) # PNGループ用にシート名を取得

        for sheet_hl_obj in workbook_hl: # openpyxlのイテレータでシートオブジェクトを取得
            try:
                dimension_hl = sheet_hl_obj.calculate_dimension()
                if dimension_hl:
                    sheet_hl_obj.print_area = dimension_hl
                    sheet_hl_obj.page_setup.fitToPage = True
                    logger.info(f"ハイライト済みシート '{sheet_hl_obj.title}' の印刷範囲を {dimension_hl} に設定しました")
            except Exception as e_dim_hl:
                logger.warning(f"ハイライト済みシート '{sheet_hl_obj.title}' の印刷範囲設定エラー: {e_dim_hl}")
        
        workbook_hl.save(highlighted_excel_path_str) # 変更をハイライト済みファイルに保存
        
        # LibreOfficeを使用してPNGに変換
        # highlighted_excel = state["highlighted_excel"] # highlighted_excel_path_str を使用
        command = f"soffice --headless --convert-to png \"{highlighted_excel_path_str}\" --outdir \"{str(captures_dir)}\""
        
        logger.info(f"実行コマンド: {command}")
        subprocess.run(command, shell=True, check=True)
        
        # 生成されたPNGファイルのパスを取得
        excel_filename = os.path.basename(highlighted_excel_path_str)
        excel_basename = os.path.splitext(excel_filename)[0]
        
        highlighted_captures = []

        if len(sheet_names_for_loop) == 1:
            # シートが1枚の場合
            capture_path = captures_dir / f"{excel_basename}.png"
            if capture_path.exists():
                highlighted_captures.append(str(capture_path))
                logger.info(f"単一シートのキャプチャファイルを発見: {capture_path}")
            else:
                # soffice がシート名を付与する場合も考慮 (例: excel_basename_Sheet1.png)
                capture_path_with_sheet_name = captures_dir / f"{excel_basename}_{sheet_names_for_loop[0]}.png"
                capture_path_with_sheet_index = captures_dir / f"{excel_basename}_sheet1.png" # 1ベースのインデックス
                capture_path_with_sheet_index_0 = captures_dir / f"{excel_basename}_sheet0.png" # 0ベースのインデックス
                
                if capture_path_with_sheet_name.exists():
                    highlighted_captures.append(str(capture_path_with_sheet_name))
                    logger.info(f"単一シートのキャプチャファイルを発見 (シート名付き): {capture_path_with_sheet_name}")
                elif capture_path_with_sheet_index.exists():
                    highlighted_captures.append(str(capture_path_with_sheet_index))
                    logger.info(f"単一シートのキャプチャファイルを発見 (シートインデックス付き): {capture_path_with_sheet_index}")
                elif capture_path_with_sheet_index_0.exists():
                    highlighted_captures.append(str(capture_path_with_sheet_index_0))
                    logger.info(f"単一シートのキャプチャファイルを発見 (シートインデックス0付き): {capture_path_with_sheet_index_0}")
                else:
                    logger.error(f"単一シートのキャプチャファイルが見つかりません。期待されたパス: {capture_path} または {capture_path_with_sheet_name} や {capture_path_with_sheet_index}")
        else:
            # シートが複数枚の場合
            for sheet_idx, sheet_name in enumerate(sheet_names_for_loop, 1):
                # LibreOfficeが出力する可能性のあるファイル名パターン
                # パターン1: <basename>_sheet<N>.png (Nは1から始まる)
                capture_path_p1 = captures_dir / f"{excel_basename}_sheet{sheet_idx}.png"
                # パターン2: <basename>-<N>.png (Nは1から始まる)
                capture_path_p2 = captures_dir / f"{excel_basename}-{sheet_idx}.png"
                # パターン3: <basename>_<シート名>.png
                capture_path_p3 = captures_dir / f"{excel_basename}_{sheet_name}.png"
                # パターン4: <basename>.png (LibreOfficeのバージョンによっては最初のシートのみこの名前になる可能性も稀にある)
                # このパターンは単一シートの場合に主に処理されるが、複数シートの最初のシートで発生する可能性も考慮するなら、ここでチェックも可能
                # capture_path_p4 = captures_dir / f"{excel_basename}.png"

                actual_capture_path = None
                if capture_path_p1.exists():
                    actual_capture_path = capture_path_p1
                elif capture_path_p2.exists():
                    actual_capture_path = capture_path_p2
                elif capture_path_p3.exists():
                    actual_capture_path = capture_path_p3
                # elif sheet_idx == 1 and capture_path_p4.exists(): # もし最初のシートがbasename.pngになる場合
                #     actual_capture_path = capture_path_p4

                if actual_capture_path:
                    highlighted_captures.append(str(actual_capture_path))
                    logger.info(f"シート '{sheet_name}' のキャプチャファイルを発見: {actual_capture_path}")
                else:
                    logger.warning(f"シート '{sheet_name}' (インデックス {sheet_idx}) のキャプチャファイルが見つかりません。試行したパターン: {capture_path_p1}, {capture_path_p2}, {capture_path_p3}")

        if not highlighted_captures and sheet_names_for_loop:
            logger.error(f"ハイライト済みExcelのキャプチャファイルが一つも生成されませんでした。Sofficeのコマンド出力を確認してください。コマンド: {command}")
            
        logger.info(f"ハイライト済みExcelキャプチャ完了: {highlighted_captures}")
        
        # 状態の更新
        return {
            **state,
            "highlighted_captures": highlighted_captures,
            "status": "進行中"
        }
        
    except Exception as e:
        logger.error(f"ハイライト済みExcelキャプチャ取得エラー: {str(e)}")
        return {
            **state,
            "status": "エラー",
            "error_message": f"ハイライト済みExcelキャプチャ取得エラー: {str(e)}"
        }

# 5. マルチモーダルLLMによる検証（structured_output使用）
def validate_with_multimodal_llm(state: ExcelFormState) -> ExcelFormState:
    """
    マルチモーダルLLMを使用してハイライト済み入力欄の検証を行う
    structured_outputを使用して確実に指定の形式で結果を受け取る
    """
    logger.info(f"マルチモーダルLLMによる検証開始 (v{state['current_iteration']})")
    
    try:
        # 実際の保存先ベースディレクトリを決定
        user_defined_output_dir = state.get("output_dir")
        if user_defined_output_dir and str(user_defined_output_dir).strip():
            base_save_path = Path(user_defined_output_dir)
        else:
            base_save_path = Path(state["excel_file"]).parent
        
        final_output_dir = base_save_path / "format_data"
        final_output_dir.mkdir(exist_ok=True, parents=True)
        
        # マルチモーダルLLMクライアントの初期化（structured_output使用）
        llm = ChatOpenAI(
            model="gpt-4.1-mini",
            temperature=0
        ).with_structured_output(ValidationResult)
        
        validation_results = []
        structured_validations = []
        
        for capture_path in state["highlighted_captures"]:
            # 画像をbase64エンコード
            with open(capture_path, "rb") as img_file:
                base64_image = base64.b64encode(img_file.read()).decode("utf-8")
            
            # 画像をbase64エンコード
            with open(state["original_excel_capture"], "rb") as img_file_original:
                base64_image_original = base64.b64encode(img_file_original.read()).decode("utf-8")

            # プロンプトの作成
            prompt = f"""
以下は、Excelフォームの画像と、入力欄として推定されたセルをハイライト（yellow）した画像です。

このハイライトされた箇所について、以下の観点で評価を行ってください。
- 入力欄として適切なセルがハイライトされているか
- 入力すべきでない欄がハイライトされていないか

問題がなければステータスを「OK」としてください。
問題がある場合は、ステータスを「修正が必要」とし、具体的な問題点と修正案を説明してください。
"""
            
            # マルチモーダルLLMに問い合わせ
            response = llm.invoke([
                HumanMessage(content=[
                    {"type": "text", "text": prompt},
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:image/png;base64,{base64_image_original}"
                        }
                    },
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:image/png;base64,{base64_image}"
                        }
                    }
                ])
            ])
            
            # 構造化された検証結果を取得
            structured_validation = response
            structured_validations.append(structured_validation)
            
            # 従来の形式のテキスト応答も生成（互換性のため）
            validation_text = f"検証結果: {structured_validation.status}\n"
            if structured_validation.issues:
                validation_text += "問題点:\n" + "\n".join([f"- {issue}" for issue in structured_validation.issues]) + "\n"
            if structured_validation.suggestions:
                validation_text += "修正案:\n" + "\n".join([f"- {suggestion}" for suggestion in structured_validation.suggestions]) + "\n"
            
            validation_results.append(validation_text)
            
            # 検証結果をログに記録
            capture_filename = os.path.basename(capture_path)
            logger.info(f"検証結果 ({capture_filename}): {structured_validation.status}")
        
        # 検証結果をファイルに保存
        validation_result_file = final_output_dir / f"validation_result_v{state['current_iteration']}.txt"
        with open(validation_result_file, "w", encoding="utf-8") as f:
            f.write(f"シート {state['current_iteration']} の検証結果:\n")
            f.write("\n\n".join(validation_results))
        logger.info(f"検証結果ファイル保存: {validation_result_file}")
        
        # 構造化された検証結果を保存
        structured_validation_file = final_output_dir / f"structured_validation_v{state['current_iteration']}.json"
        with open(structured_validation_file, "w", encoding="utf-8") as f:
            # 複数の検証結果がある場合は最初のものを使用
            f.write(structured_validations[0].model_dump_json(indent=2))
        logger.info(f"構造化検証結果ファイル保存: {structured_validation_file}")
        
        # 検証結果の分析
        validation_status = "OK"
        if any(validation.status == "修正が必要" for validation in structured_validations):
            validation_status = "修正が必要"
        
        logger.info(f"検証完了: 結果={validation_status}")
        
        # 状態の更新
        return {
            **state,
            "validation_result": "\n\n".join(validation_results),
            "structured_validation": structured_validations[0],  # 複数ある場合は最初のものを使用
            "validation_status": validation_status,
            "status": "進行中"
        }
        
    except Exception as e:
        logger.error(f"検証エラー: {str(e)}")
        return {
            **state,
            "status": "エラー",
            "error_message": f"検証エラー: {str(e)}"
        }

# 6. 入力欄情報の修正（structured_output使用）
def correct_fields_with_multimodal_llm(state: ExcelFormState) -> ExcelFormState:
    """
    検証結果に基づいて入力欄情報を修正する
    structured_outputを使用して確実に指定の形式で結果を受け取る
    """
    logger.info(f"入力欄情報の修正開始 (v{state['current_iteration'] + 1})")
    
    try:
        # 実際の保存先ベースディレクトリを決定
        user_defined_output_dir = state.get("output_dir")
        if user_defined_output_dir and str(user_defined_output_dir).strip():
            base_save_path = Path(user_defined_output_dir)
        else:
            base_save_path = Path(state["excel_file"]).parent
        
        final_output_dir = base_save_path / "format_data"
        final_output_dir.mkdir(exist_ok=True, parents=True)
        
        # 現在の推定結果
        structured_fields = state["structured_fields"]
        
        # 検証結果
        structured_validation = state["structured_validation"]
        
        # ハイライトされたExcel画像
        with open(state["highlighted_captures"][0], "rb") as img_file:
            base64_image = base64.b64encode(img_file.read()).decode("utf-8")
        
        # 元のExcelフォームの画像
        with open(state["original_excel_capture"], "rb") as img_file_original:
            base64_image_original = base64.b64encode(img_file_original.read()).decode("utf-8")

        # マルチモーダルLLMクライアントの初期化（structured_output使用）
        llm = ChatOpenAI(
            model="gpt-4.1-mini",
            temperature=0
        ).with_structured_output(CollectExcelFormFields)
        
        # プロンプトの作成
        prompt = f"""
あなたはExcelフォームの入力欄を特定する専門家です。
以下のSTEPで作業をしてください。
STEP1:以下の情報をよく確認してください。
- 現在推定されているExcelフォームの入力欄情報
{structured_fields.model_dump_json(indent=2)}

- 添付の画像 ※元のExcelフォームの画像と、推定された入力欄をハイライトしたExcelシートの画像

- これらに対するレビュー結果
{structured_validation.model_dump_json(indent=2)}

STEP2:検証結果と画像に基づいて、修正すべき箇所を回答してください。
"""
        
        # マルチモーダルLLMに問い合わせ
        response = llm.invoke([
            HumanMessage(content=[
                {"type": "text", "text": prompt},
                {
                    "type": "image_url",
                    "image_url": {
                        "url": f"data:image/png;base64,{base64_image_original}"
                    }
                },
                {
                    "type": "image_url",
                    "image_url": {
                        "url": f"data:image/png;base64,{base64_image}"
                    }
                }
            ])
        ])
        
        # 構造化された応答を取得
        correction_instructions = response  # CollectExcelFormFields 型

        # 現在のフィールドリストを取得
        current_fields_list = list(state["structured_fields"].fields)
        current_fields_dict = {field.cell_id: field for field in current_fields_list}

        # 削除するフィールドを処理
        for field_to_delete in correction_instructions.delete_fields:
            if field_to_delete.cell_id in current_fields_dict:
                del current_fields_dict[field_to_delete.cell_id]

        # 追加するフィールドを処理
        for field_to_add in correction_instructions.add_fields:
            current_fields_dict[field_to_add.cell_id] = field_to_add
        
        updated_fields_list = list(current_fields_dict.values())

        # 更新されたExcelFormFieldsを作成
        updated_structured_fields = ExcelFormFields(
            fields=updated_fields_list,
            reason=correction_instructions.reason  # LLMからの判断根拠を使用
        )
        
        # 従来の形式（Dict[str, str]）に変換（互換性のため）
        corrected_fields = {}
        for field in updated_structured_fields.fields:
            corrected_fields[field.cell_id] = field.description
        
        next_iteration = state["current_iteration"] + 1
        
        structured_fields_file = final_output_dir / f"structured_fields_v{next_iteration}.json"
        with open(structured_fields_file, "w", encoding="utf-8") as f:
            f.write(updated_structured_fields.model_dump_json(indent=2))
        
        # 従来の形式も保存（互換性のため）
        corrected_fields_file = final_output_dir / f"estimated_fields_v{next_iteration}.json"
        with open(corrected_fields_file, "w", encoding="utf-8") as f:
            json.dump(corrected_fields, f, ensure_ascii=False, indent=2)
        
        logger.info(f"入力欄情報の修正完了: {structured_fields_file}")
        
        # 状態の更新
        return {
            **state,
            "estimated_fields": corrected_fields,
            "structured_fields": updated_structured_fields, # 更新された情報をセット
            "current_iteration": next_iteration,
            "status": "進行中"
        }
        
    except Exception as e:
        logger.error(f"入力欄情報の修正エラー: {str(e)}")
        return {
            **state,
            "status": "エラー",
            "error_message": f"入力欄情報の修正エラー: {str(e)}"
        }

# 7. 最終結果の生成
def generate_final_json(state: ExcelFormState) -> ExcelFormState:
    """
    最終的な入力欄情報JSONを生成する
    """
    logger.info("最終結果の生成")
    
    try:
        # 実際の保存先ベースディレクトリを決定
        user_defined_output_dir = state.get("output_dir")
        if user_defined_output_dir and str(user_defined_output_dir).strip():
            base_save_path = Path(user_defined_output_dir)
        else:
            base_save_path = Path(state["excel_file"]).parent
        
        final_output_dir = base_save_path / "format_data"
        final_output_dir.mkdir(exist_ok=True, parents=True)
        
        # 最終的な入力欄情報
        final_fields = state["estimated_fields"]
        final_structured_fields = state["structured_fields"]
        
        # 結果をファイルに保存
        final_json_file = final_output_dir / "final_form_definition.json"
        with open(final_json_file, "w", encoding="utf-8") as f:
            json.dump(final_fields, f, ensure_ascii=False, indent=2)
        
        # 構造化された形式も保存
        final_structured_file = final_output_dir / "final_structured_form_definition.json"
        with open(final_structured_file, "w", encoding="utf-8") as f:
            f.write(final_structured_fields.model_dump_json(indent=2))
        
        logger.info(f"処理が完了しました。最終結果: {final_json_file}")
        
        # 状態の更新
        return {
            **state,
            "final_json": str(final_json_file),
            "status": "完了"
        }
        
    except Exception as e:
        logger.error(f"最終結果の生成エラー: {str(e)}")
        return {
            **state,
            "status": "エラー",
            "error_message": f"最終結果の生成エラー: {str(e)}"
        }

# ルーター関数: 検証結果に基づいて次のステップを決定
def router(state: ExcelFormState) -> str:
    """
    状態に基づいて次のステップを決定する
    """
    # エラーが発生した場合は終了
    if state["status"] == "エラー":
        return END
    
    # 処理が完了した場合は終了
    if state["status"] == "完了":
        return END
    
    # 検証結果に基づく分岐
    if state["validation_status"] == "OK":
        return "generate_final_json"
    
    # 最大反復回数に達した場合は最終結果を生成
    if state["current_iteration"] >= state["max_iterations"]:
        logger.warning(f"最大反復回数 ({state['max_iterations']}) に到達しました")
        return "generate_final_json"
    
    # 修正が必要な場合は修正ステップへ
    return "correct_fields_with_multimodal_llm"

# LangGraphワークフローの構築
def build_workflow() -> StateGraph:
    """
    Excel入力欄特定ワークフローを構築する
    """
    # グラフの作成
    workflow = StateGraph(ExcelFormState)
    
    # ノードの追加
    workflow.add_node("extract_excel_data_and_capture", extract_excel_data_and_capture)
    workflow.add_node("estimate_fields_with_multimodal_llm", estimate_fields_with_multimodal_llm)
    workflow.add_node("highlight_fields", highlight_fields)
    workflow.add_node("capture_highlighted_excel", capture_highlighted_excel)
    workflow.add_node("validate_with_multimodal_llm", validate_with_multimodal_llm)
    workflow.add_node("correct_fields_with_multimodal_llm", correct_fields_with_multimodal_llm)
    workflow.add_node("generate_final_json", generate_final_json)
    
    # エッジの追加（基本フロー）
    workflow.add_edge("extract_excel_data_and_capture", "estimate_fields_with_multimodal_llm")
    workflow.add_edge("estimate_fields_with_multimodal_llm", "highlight_fields")
    workflow.add_edge("highlight_fields", "capture_highlighted_excel")
    workflow.add_edge("capture_highlighted_excel", "validate_with_multimodal_llm")
    
    # 条件分岐
    workflow.add_conditional_edges(
        "validate_with_multimodal_llm",
        router,
        {
            "correct_fields_with_multimodal_llm": "correct_fields_with_multimodal_llm",
            "generate_final_json": "generate_final_json",
            END: END
        }
    )
    
    # 修正後のフロー
    workflow.add_edge("correct_fields_with_multimodal_llm", "highlight_fields")
    
    # 開始ノードの設定
    workflow.set_entry_point("extract_excel_data_and_capture")
    
    return workflow

